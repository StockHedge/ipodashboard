#!/usr/bin/env python3
"""
IPO 공모주 데이터 추출 스크립트
- 사용자의 xlsx (공모주데이터 시트)에서 최근 N년 데이터만 추출
- 대시보드용 깨끗한 JSON 배열로 출력
- 한글 경로 안전 (pathlib + 사용자 환경 기준)

사용법 (PowerShell 예):
  python scripts/extract_ipo_data.py --xlsx "$env:USERPROFILE\Downloads\새 계정 프로그램 시트 (1).xlsx" --years 3 --out ../data/ipo-recent.json

필수: pandas, openpyxl (또는 openpyxl만으로도 동작)
"""

from __future__ import annotations
import argparse
import json
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from openpyxl import load_workbook
except ImportError:
    print("오류: openpyxl이 필요합니다. pip install openpyxl pandas", file=sys.stderr)
    sys.exit(1)


def find_main_sheet(wb) -> str:
    """'공모주데이터'가 포함된 주요 데이터 시트 찾기"""
    for name in wb.sheetnames:
        if "공모주데이터" in name and "팩" not in name and "확인" not in name:
            return name
    # fallback: 가장 큰 행 수를 가진 시트
    return max(wb.sheetnames, key=lambda n: wb[n].max_row)


def _normalize_header(h: Any) -> str:
    """헤더 문자열에서 공백/줄바꿈/(억) 같은 단위 표기 제거 → 키워드 매칭 일관성 확보.
    예: '확정\n공모\n금액\n(억)' → '확정공모금액'."""
    if h is None:
        return ""
    s = str(h).lower()
    # 모든 whitespace 제거 (\n, \t, 공백 등)
    s = re.sub(r"\s+", "", s)
    # (n:1), (억), (%) 같은 후행 단위 표기 제거
    s = re.sub(r"\([^)]*\)", "", s)
    return s


def guess_column_indices(headers: list[Any]) -> dict[str, int | None]:
    """
    헤더 이름으로 주요 컬럼 인덱스 추정 (0-based).
    - whitespace 정규화 (`\\n` 헤더 대응)
    - used-tracking 으로 동일 인덱스가 두 필드에 매핑되는 충돌 차단
    - 사용자 시트 ("종목명, 담당자, 증권신고서, 상장일, 종목코드, 주관사, ...") 실측 헤더 기반 키워드
    """
    h_norm = [_normalize_header(h) for h in headers]
    used: set[int] = set()

    def find(*keywords: str, reusable: bool = False) -> int | None:
        for kw in keywords:
            kw_norm = _normalize_header(kw)
            for i, hh in enumerate(h_norm):
                if (not reusable) and i in used:
                    continue
                if kw_norm and kw_norm in hh:
                    if not reusable:
                        used.add(i)
                    return i
        return None

    mapping: dict[str, int | None] = {}
    mapping["name"] = find("종목명")
    mapping["listing_date"] = find("상장일")
    # 종목코드 컬럼 (예: 'KOSDAQ:302550') — ticker + market 추출에 모두 사용 (reusable)
    mapping["stock_code"] = find("종목코드", reusable=True)
    mapping["market"] = find("시장", "거래소", "종목코드", reusable=True)
    mapping["sector"] = find("섹터", "업종", "산업분류")
    mapping["underwriter"] = find("주관사", "대표주관", "주간사")
    # 확정 공모 금액 (억) — 시가총액(인덱스 14)과 분명히 분리
    mapping["offering_amount"] = find("확정공모금액", "공모금액", "공모총액")
    mapping["offering_price"] = find("확정공모가", "공모가", "발행가")
    # 수요예측 밴드 (#5 모듈: 밴드 vs 확정가 산점도용)
    mapping["band_low"] = find("밴드하단", "밴드 하단", "공모가밴드하단")
    mapping["band_high"] = find("밴드상단", "밴드 상단", "공모가밴드상단")
    # 시초매도/종가매도/고가매도 명확 구분 — '첫날 수익률' = 시초매도 기본, 종가매도는 별도 필드
    mapping["first_day_return"] = find("시초매도시수익률", "시초매도", "시초")
    mapping["first_day_close_return"] = find("종가매도시수익률", "종가매도")
    mapping["high_return"] = find("상장일고가매도시수익률", "상장일고가매도", "고가매도시수익률", "고가매도")
    # 고가형성시간 (실제로 '시간' 단위 — 시간/분/초). days 단위가 시트에 없으므로 None 허용.
    mapping["days_to_high"] = find("고가형성일", "고가달성일", "달성일", "달성기간")
    mapping["high_time"] = find("상장일고가형성시간", "고가형성시간", "형성시간")
    # 1개월/3개월/6개월 — '매도시수익률' 명시 (단순 '1개월'은 종가 컬럼과 충돌)
    mapping["return_1m"] = find("1개월매도시수익률", "1개월매도수익률", "1개월매도")
    mapping["return_3m"] = find("3개월매도시수익률", "3개월매도수익률", "3개월매도")
    mapping["return_6m"] = find("6개월매도시수익률", "6개월매도수익률", "6개월매도")
    mapping["competition_retail"] = find("개인청약경쟁률", "개인경쟁률", "청약경쟁률")
    mapping["competition_inst"] = find("기관수요예측경쟁률", "기관경쟁률", "수요예측경쟁률")
    # 의무보유확약 = 락업 비율 (한국 IPO 도메인 핵심 시그널)
    mapping["lockup_rate"] = find("락업비율", "참여기관락업비율", "의무보유확약", "확약비율")
    # 상한가/따따상 컬럼은 사용자 시트에 미존재 — high_return >= 30 에서 client-side 계산
    mapping["upper_limit"] = find("따따상", "따상", "+30", "상한가")
    return mapping


# 대분류 매핑 (사용자 시트의 raw 섹터 → UI 일관성 위한 대분류 group)
# 더 구체적인 키워드를 먼저 (예: '2차전지' 이 '전지' 보다 먼저) 평가하여 정확도 확보
SECTOR_GROUPS: list[tuple[list[str], str]] = [
    (["스팩", "spac"], "스팩"),
    (["2차전지", "이차전지", "배터리"], "2차전지/소재"),
    (["반도체"], "반도체/장비"),
    (["디스플레이", "ole d"], "디스플레이"),
    (["의료기", "헬스케어"], "의료기기"),
    (["바이오", "제약", "신약"], "바이오/제약"),
    (["뷰티", "화장품", "코스메"], "화장품/소비재"),
    (["게임"], "게임/콘텐츠"),
    (["엔터", "콘텐츠", "미디어"], "엔터/콘텐츠"),
    (["sw", "s/w", "소프트웨어", "ai", "인공지능", "saas", "플랫폼", "솔루션", "it"], "IT/소프트웨어"),
    (["로봇", "로보"], "로봇/AI"),
    (["자동차", "모빌리티", "전기차"], "자동차/모빌리티"),
    (["조선", "해운"], "조선/해운"),
    (["식품", "음료", "외식"], "식품/음료"),
    (["신재생", "친환경", "에너지", "수소", "태양광"], "신재생에너지"),
    (["핀테크", "금융", "증권", "은행", "보험"], "금융/핀테크"),
    (["화학", "소재"], "화학/소재"),
    (["기계", "장비", "부품"], "기계/장비"),
    (["반도체장비"], "반도체/장비"),
    (["건설", "건자재"], "건설/자재"),
    (["통신", "네트워크"], "통신/네트워크"),
]


def normalize_sector(raw: Any) -> str:
    """raw 섹터 문자열 → UI 대분류. 매치 없으면 '기타'."""
    if raw is None:
        return "기타"
    s = str(raw).strip().lower()
    if not s or s == "nan":
        return "기타"
    for keywords, group in SECTOR_GROUPS:
        for kw in keywords:
            if kw.lower() in s:
                return group
    return "기타"


def parse_stock_code(value: Any) -> tuple[str, str | None]:
    """
    종목코드 컬럼 값에서 (market, ticker) 추출. ticker 추출 실패 시 None.
    지원 패턴:
      'KOSDAQ:302550', 'KOSPI:005930', 'KS:005930', 'KSE:005930',
      '유가증권:005930', '코스피:005930', '코스닥:302550',
      단순 6자리 '005930' (선두 0 보존),
      Yahoo 스타일 '005930.KS' / '302550.KQ'
    KOSPI 룰: 6자리 ticker 가 '0' 으로 시작하고 prefix 단서 없으면 KOSPI 가능성 ↑
    그러나 한국 거래소 실제 룰은 더 복잡 → prefix 단서 우선, 단서 없으면 KOSDAQ default.
    """
    if value is None:
        return ("KOSDAQ", None)
    s = str(value).strip().upper()
    if not s or s == "NAN":
        return ("KOSDAQ", None)
    # 한글 prefix 정규화 (대소문자 영향 없음)
    raw = str(value).strip()
    market = "KOSDAQ"
    # KOSPI 단서 (영문/한글/Yahoo suffix)
    kospi_hints = ("KOSPI", "KS:", "KSE:", "유가", "코스피", ".KS")
    kosdaq_hints = ("KOSDAQ", "KQ:", "KQE:", "코스닥", ".KQ")
    if any(kw in s for kw in kospi_hints) or any(kw in raw for kw in kospi_hints):
        market = "KOSPI"
    elif any(kw in s for kw in kosdaq_hints) or any(kw in raw for kw in kosdaq_hints):
        market = "KOSDAQ"
    digits = re.sub(r"\D", "", s)
    ticker = digits[-6:] if len(digits) >= 6 else None
    return (market, ticker)


# 소수 형태(0.0846) vs 퍼센트(8.46) 판별 임계값.
# 사용자 시트는 소수 형태가 일반적. |v| < 5 면 소수로 간주 (×100).
# 경계 주의: 실제 +4.9% 수익률이 소수 4.9 로 저장된 경우 오분류 가능 →
# 시트 포맷이 일관(소수)이라는 전제. 혼재 시 openpyxl number_format 기반 판별로 교체 권장.
PERCENT_DECIMAL_THRESHOLD = 5.0

def to_percent(value: Any) -> float | None:
    if value is None:
        return None
    try:
        v = float(value)
    except (ValueError, TypeError):
        return None
    if v != v:  # NaN
        return None
    if -PERCENT_DECIMAL_THRESHOLD < v < PERCENT_DECIMAL_THRESHOLD:
        return v * 100.0
    return v


def parse_listing_date(value: Any) -> str | None:
    """
    listingDate 를 'YYYY-MM-DD' 문자열로 표준화.
    지원 형식: datetime 객체, 'YYYY-MM-DD', 'YYYY.MM.DD', 'YYYY/MM/DD', 'YYYYMMDD'.
    실패 시 None 반환 (호출 측에서 row skip).
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return None
    # 이미 표준 형식
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    # 점/슬래시 구분자
    for sep in (".", "/"):
        if sep in s and len(s.split(sep)) == 3:
            try:
                y, m, d = s.split(sep)
                return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
            except (ValueError, TypeError):
                continue
    # YYYYMMDD
    if len(s) == 8 and s.isdigit():
        return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
    return None


def safe_cell(row: list[Any], col_map: dict[str, int | None], key: str, default_idx: int | None, default_val: Any = None) -> Any:
    """
    col_map[key] 가 None 이거나 누락 → default_idx 사용.
    default_idx 도 None 이면 default_val 반환.
    row 인덱스가 범위 밖이어도 default_val 반환.

    이전 버그: col_map.get(key, default_idx) 는 key 가 있고 값이 None 인 경우 None 을 반환
    → row[None] → TypeError. 이 helper 가 그 경로를 차단한다.
    """
    idx = col_map.get(key)
    if idx is None:
        idx = default_idx
    if idx is None:
        return default_val
    try:
        return row[idx]
    except (IndexError, TypeError):
        return default_val


def safe_float(v: Any, default: float | None = None) -> float | None:
    try:
        if v is None or str(v).lower() in ("", "nan", "none"):
            return default
        return float(v)
    except (ValueError, TypeError):
        return default


def extract_row(row: list[Any], col_map: dict[str, int | None], headers: list[Any]) -> dict | None:
    """한 행을 대시보드 스키마로 변환 (None-safe + 사용자 시트 실측 헤더 기반)"""
    try:
        name = str(safe_cell(row, col_map, "name", 0, "") or "").strip()
        if not name or name.lower() == "nan":
            return None

        listing_date = parse_listing_date(safe_cell(row, col_map, "listing_date", 8, None))
        if listing_date is None:
            return None

        # 종목코드 컬럼에서 시장(KOSPI/KOSDAQ) + ticker(6자리) 동시 추출
        market, ticker = parse_stock_code(safe_cell(row, col_map, "stock_code", None, None))
        # 별도 market 컬럼이 있으면 보강
        market_val = safe_cell(row, col_map, "market", None, None)
        if market_val and "KOSPI" in str(market_val).upper():
            market = "KOSPI"

        sector_raw = str(safe_cell(row, col_map, "sector", None, "기타") or "기타").strip()[:30]
        sector = normalize_sector(sector_raw)  # UI 대분류 (필터 일관성)

        # 수익률은 소수 형태 (0.0846) → 퍼센트 (8.46) 로 정규화
        first_day = to_percent(safe_cell(row, col_map, "first_day_return", None))
        first_day_close = to_percent(safe_cell(row, col_map, "first_day_close_return", None))
        high_ret = to_percent(safe_cell(row, col_map, "high_return", None))
        return_1m = to_percent(safe_cell(row, col_map, "return_1m", None))
        return_3m = to_percent(safe_cell(row, col_map, "return_3m", None))
        return_6m = to_percent(safe_cell(row, col_map, "return_6m", None))

        # 락업 비율 (의무보유확약) — 소수 또는 % 가능. 0~1 범위면 ×100
        lockup_raw = safe_float(safe_cell(row, col_map, "lockup_rate", None), None)
        lockup_pct = None
        if lockup_raw is not None:
            lockup_pct = lockup_raw * 100.0 if -0.001 <= lockup_raw <= 1.001 else lockup_raw

        # 상한가/따상 컬럼이 시트에 없으므로 high_return >= 30 으로 client-side 추정
        # (한국 상한가 +30% — IPO 첫날에는 의미상 따상에 가장 가까운 시그널)
        upper_limit_explicit = safe_cell(row, col_map, "upper_limit", None, None)
        if upper_limit_explicit is not None:
            upper_limit_hit = bool(upper_limit_explicit)
        else:
            upper_limit_hit = bool(high_ret is not None and high_ret >= 30.0)

        return {
            "id": hash(name + listing_date) % 1_000_000,  # 간단 ID
            "name": name,
            "ticker": ticker,
            "listingDate": listing_date,
            "market": market,
            "sector": sector,           # 대분류 (필터·집계용)
            "sectorRaw": sector_raw,    # 시트 원본 (참고용, UI 상세 모달에 표시 가능)
            "offeringPrice": int(safe_float(safe_cell(row, col_map, "offering_price", None), 10000) or 10000),
            "bandLow": safe_float(safe_cell(row, col_map, "band_low", None), None),
            "bandHigh": safe_float(safe_cell(row, col_map, "band_high", None), None),
            "firstDayReturn": first_day if first_day is not None else 0.0,
            "firstDayCloseReturn": first_day_close,
            "highReturn": high_ret if high_ret is not None else 0.0,
            # 사용자 시트는 '상장일고가매도시 수익률' 만 보유 → 의미상 daysToHigh = 1 (상장 당일).
            # 일자 단위 컬럼이 시트에 추가되면 자동으로 그 값을 사용.
            "daysToHigh": int(safe_float(safe_cell(row, col_map, "days_to_high", None), 1) or 1),
            "return1M": return_1m,
            "return3M": return_3m,
            "return6M": return_6m,
            "competitionRetail": safe_float(safe_cell(row, col_map, "competition_retail", None), 50.0) or 50.0,
            "competitionInst": safe_float(safe_cell(row, col_map, "competition_inst", None), 15.0) or 15.0,
            "offeringAmount": int(safe_float(safe_cell(row, col_map, "offering_amount", None), 500) or 500),
            "lockupRate": lockup_pct,
            "upperLimitHit": upper_limit_hit,
            "underwriter": str(safe_cell(row, col_map, "underwriter", None, "주간사") or "주간사")[:20],
        }
    except (KeyError, ValueError, TypeError, AttributeError, IndexError) as e:
        print(f"  [skip] extract_row fail: {e.__class__.__name__}: {e}", file=sys.stderr)
        return None


def main():
    parser = argparse.ArgumentParser(description="IPO xlsx → 대시보드 JSON 추출")
    parser.add_argument("--xlsx", required=True, help="원본 xlsx 파일 경로")
    parser.add_argument("--years", type=int, default=3, help="최근 N년 (기본 3)")
    parser.add_argument("--out", required=True, help="출력 JSON 경로")
    parser.add_argument(
        "--debug-headers",
        action="store_true",
        help="추출 실패 디버그용: 시트 헤더와 첫 1개 데이터 행을 전부 출력",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx).expanduser().resolve()
    out_path = Path(args.out).expanduser().resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    if not xlsx_path.exists():
        print(f"오류: 파일을 찾을 수 없습니다: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    print(f"파일 로드 중: {xlsx_path.name}")

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet_name = find_main_sheet(wb)
    ws = wb[sheet_name]
    print(f"데이터 시트: {sheet_name} (총 {ws.max_row or '?'}행)")

    # 헤더 (1행)
    headers = [cell.value for cell in ws[1]]
    col_map = guess_column_indices(headers)

    # 매핑 결과를 None 포함하여 명확히 출력 (사용자가 어떤 키가 매핑 실패했는지 즉시 인지)
    print("추정 컬럼 매핑:")
    for k, v in col_map.items():
        flag = "" if v is not None else "  ⚠️ 매핑 실패 (해당 필드는 default 값으로 대체)"
        print(f"  {k:22s}: {v}{flag}")

    if args.debug_headers:
        print(f"\n=== 헤더 전체 dump ({len(headers)}개) ===")
        for i, h in enumerate(headers):
            print(f"  [{i:2d}] {h!r}")
        # 첫 데이터 행 미리보기
        try:
            first_row = next(ws.iter_rows(min_row=2, values_only=True))
            print(f"\n=== 첫 데이터 행 미리보기 ({len(first_row)}컬럼) ===")
            for i, v in enumerate(first_row):
                print(f"  [{i:2d}] {v!r}")
        except StopIteration:
            print("  (데이터 행 없음)")

    # 날짜 필터
    cutoff = datetime.now() - timedelta(days=args.years * 365 + 30)
    records = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row_idx % 500 == 0:
            print(f"  처리 중... {row_idx}행")

        rec = extract_row(list(row), col_map, headers)
        if not rec:
            continue

        try:
            listing_dt = datetime.strptime(rec["listingDate"], "%Y-%m-%d")
            if listing_dt >= cutoff:
                records.append(rec)
        except (ValueError, KeyError, TypeError) as e:
            # 도메인 의미 기반: listingDate 형식이 안 맞거나 누락 → 통계에 포함하지 않음
            print(
                f"  [skip] row {row_idx}: listingDate parse fail "
                f"(name={rec.get('name')!r}, listingDate={rec.get('listingDate')!r}, err={e.__class__.__name__})",
                file=sys.stderr,
            )

    wb.close()

    # ticker 누락 종목 fdr 보완 — 시트 [9] 종목코드가 비거나 형식 다른 경우 종목명으로 역조회
    missing_ticker = [r for r in records if not r.get("ticker") and r.get("name")]
    if missing_ticker:
        try:
            import FinanceDataReader as fdr
            krx = fdr.StockListing("KRX")[["Code", "Name"]]
            name_to_code = {}
            for _, row in krx.iterrows():
                name_to_code[str(row["Name"]).strip()] = str(row["Code"]).zfill(6)
            fixed = 0
            for r in missing_ticker:
                code = name_to_code.get(str(r["name"]).strip())
                if code:
                    r["ticker"] = code
                    fixed += 1
            if fixed:
                print(f"fdr 보완: {fixed}/{len(missing_ticker)}건 ticker 자동 추가")
        except ImportError:
            print(f"fdr 미설치 — ticker 누락 {len(missing_ticker)}건 그대로 (pip install finance-datareader)", file=sys.stderr)
        except Exception as e:
            print(f"fdr 보완 실패: {e}", file=sys.stderr)

    # ID 재부여 (정렬)
    records.sort(key=lambda r: r["listingDate"], reverse=True)
    for i, r in enumerate(records, 1):
        r["id"] = i

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

    # 추가 출력: file:// 환경에서도 동작하도록 JS 변수 형태로 sibling 파일 생성
    # index.html 이 <script src="data/ipo-recent.js"> 로 동기 로드 → window.IPO_DATA_OVERRIDE 우선 사용
    js_path = out_path.with_suffix(".js")
    js_body = (
        "// Auto-generated by scripts/extract_ipo_data.py — DO NOT EDIT MANUALLY.\n"
        "// index.html 이 이 파일을 자동 로드하여 fetch(file://) 차단 환경에서도 데이터가 즉시 보임.\n"
        f"// generated_at: {datetime.now().isoformat(timespec='seconds')}\n"
        f"// records: {len(records)}\n"
        f"window.IPO_DATA_OVERRIDE = {json.dumps(records, ensure_ascii=False, indent=2)};\n"
    )
    js_path.write_text(js_body, encoding="utf-8")

    print(f"\n완료: {len(records)}건 추출 → {out_path}")
    print(f"또한 {js_path.name} 생성 — index.html 자동 로드 (file:// 도 지원)")
    if records:
        print(f"기간: {records[-1]['listingDate']} ~ {records[0]['listingDate']}")
        # 첫/끝 1건씩 미리보기 (데이터 무결성 빠른 확인용)
        print("\n첫 레코드 미리보기 (최신):")
        for k in ("name", "listingDate", "market", "sector", "offeringPrice", "firstDayReturn"):
            print(f"  {k}: {records[0].get(k)}")
        print("대시보드에서 'JSON 로드' 또는 data/ 폴더에 배치 후 사용하세요.")
    else:
        print(
            "\n⚠️ 추출된 레코드가 0건입니다. 가능한 원인:\n"
            "  1) 컬럼 매핑 실패 (위 '⚠️ 매핑 실패' 항목 확인)\n"
            "  2) listingDate 형식이 yyyy-mm-dd 또는 datetime 이 아님\n"
            "  3) cutoff (최근 N년) 기준에 들어오는 데이터가 없음\n"
            "디버그: `python scripts/extract_ipo_data.py ... --debug-headers` 로 헤더/첫행 dump 실행.",
            file=sys.stderr,
        )
        sys.exit(2)


if __name__ == "__main__":
    main()